#
# Ducumentation at https://openpyxl.readthedocs.io/en/stable/index.html
#

from openpyxl import Workbook, load_workbook

def readStatFile():
    #opens the excel file
    wb = load_workbook('Fantasy_Football_Example_stats_to_work_with.xlsx')

    playerList = {}

    #loops through worksheets (ws)
    for ws in wb:

        #saves the name of the worksheet
        positionName = ws.title
        
        #saves the first row of column names from the worksheet
        statNames = ws[1]
        
        #itterates through rows 2 and down and takes their values
        for row in ws.iter_rows(min_row=2):
            #saves the player name
            playerName = row[0].value
            #saves the players position
            playerStats = {"Postition": positionName}
            #saves the players other stats
            for index in range(1,len(row)):
                playerStats[ statNames[index].value ] = row[index].value
            #saves the players stats in playerList
            playerList[playerName] = playerStats
            
    return playerList