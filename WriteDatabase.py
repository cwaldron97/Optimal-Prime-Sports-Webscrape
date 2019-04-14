from openpyxl import Workbook, load_workbook

#Loads workbook or creates one if there is none
try:
    wb = load_workbook('Fantasy_Football_Database.xlsx')

    #assigns worksheets to variables
    for ws in wb:
        if ws.title == "QB":
            wsQB = ws
        if ws.title == "RB":
            wsRB = ws
        if ws.title == "TE":
            wsTE = ws
        if ws.title == "Defence":
            wsDefence = ws
        if ws.title == "Kicker":
            wsKicker = ws
except FileNotFoundError:
    #creates workbook and worksheets
    wb = Workbook()

    wsQB = wb.active
    wsQB.title = "QB"
    wsQB.append(["Player","TOT PTS","AVG PTS","CMP %","Pass YDS","QB RTG","TD","INT","Rush TD","Rush YDS"])

    wsRB = wb.create_sheet(title = "RB")
    wsRB.append(["Player","Rush YDS","Rush TD","AVG Carry","AVG YDS","Recieving YDS","Recieving TD","AVD REC","TOT PTS","AVG PTS"])

    wsTE = wb.create_sheet(title = "TE")
    wsTE.append(["Player","REC","Recieving YDS","Recieving TD","TOT PTS","AVG PTS"])

    wsDefence = wb.create_sheet(title = "Defence")
    wsDefence.append(["Player","TOT PTS","AVG PTS","INT","FUM","SK"])

    wsKicker = wb.create_sheet(title = "Kicker")
    wsKicker.append(["Player","TOT PTS","AVG PTS","XPM","XPA"])

    wb.save('Fantasy_Football_Database.xlsx')

#writes statistics to the worksheets
def writeToDatabase(position, statsList):
    if position == 'QB':
        wsQB.append(statsList)
    elif position == 'RB':
        wsRB.append(statsList)
    elif position == 'TE':
        wsTE.append(statsList)
    elif position == 'Defence':
        wsDefence.append(statsList)    
    elif position == 'Kicker':
        wsDefence.append(statsList)
    
    wb.save('Fantasy_Football_Database.xlsx')

def readStatFile(name):

    playerStats = None

    #loops through worksheets (ws)
    for ws in wb:

        #saves the name of the worksheet
        positionName = ws.title
        
        #saves the first row of column names from the worksheet
        statNames = ws[1]
        
        #itterates through rows 2 and down and takes their values
        for row in ws.iter_rows(min_row=2):
            #saves the player name
            playerName = row[0].value

            if (playerName == name):
                #saves the players position
                playerStats = {"Postition": positionName}
                #saves the players other stats
                for index in range(1,len(row)):
                    playerStats[ statNames[index].value ] = row[index].value
            
    return playerStats